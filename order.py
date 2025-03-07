import logging
import os
import re

import cn2an
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# 配置日志记录
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# 0: 伊花集-斯亚旦植物发皂（蓝）
# 1: 伊花集-斯亚旦植物发皂（绿）
# 2: 伊花集-奥斯曼植物发油
# 3: 伊花集-茶树植物内衣清洁皂
# 4: 伊花集-驼乳玫瑰芳华皂
# 5: 伊花集-斯亚旦·植物养发系列礼盒（2号蓝色）
# 6: 伊花集-斯亚旦草本植物|洗发皂礼盒（1号粉色）
# 7: 伊花集-起泡瓶
# 8: 伊花集-液体精油皂
# 9: 伊花集-儿童洗发皂
# 10: 伊花集-大手提袋
# 11: 伊花集-小手提袋
# 12: 伊花集-初心大号羊角刮痧板套装
# 13: 伊花集-初心整木按摩梳-皮灰黑檀
# 14: 伊花集-宣传册
golden = ('伊花集-斯亚旦植物发皂（蓝）',
          '伊花集-斯亚旦植物发皂（绿）',
          '伊花集-奥斯曼植物发油',
          '伊花集-茶树植物内衣清洁皂',
          '伊花集-驼乳玫瑰芳华皂',
          '伊花集-斯亚旦·植物养发系列礼盒（2号蓝色）',
          '伊花集-斯亚旦草本植物|洗发皂礼盒（1号粉色）',
          '伊花集-起泡瓶',
          '伊花集-液体精油皂',
          '伊花集-儿童洗发皂',
          '伊花集-大手提袋',
          '伊花集-小手提袋',
          '伊花集-初心大号羊角刮痧板套装',
          '伊花集-初心整木按摩梳-皮灰黑檀',
          '伊花集-宣传册'
          )
order_num_list = ['订单编号', '关联单号']
courier_list = ['快递公司', '物流公司']
courier_num_list = ['快递单号', '物流单号']
name_list = ['姓名', '收件人']
phone_num_list = ['电话']
address_list = ['地址']
goods_list = ['商品', '货品摘要']
quantity_list = ['数量', '数量合计']
shipper_list = ['发货人']
shipper_num_list = ['发货电话']


def ordertrans(excel_path):
    excel_file = pd.ExcelFile(excel_path)
    sheet_names = excel_file.sheet_names
    logging.info('页面列表：%s' % sheet_names)
    df = excel_file.parse(sheet_names[0])
    column_name = df.columns
    logging.info('列名列表：%s' % column_name.to_list())
    unknown_column_name = []
    new_column_name = []
    for i in range(column_name.size):
        x = column_name[i].strip()
        if x in order_num_list:
            new_column_name.append(order_num_list[0])
        elif x in courier_list:
            new_column_name.append(courier_list[0])
        elif x in courier_num_list:
            new_column_name.append(courier_num_list[0])
        elif x in name_list:
            new_column_name.append(name_list[0])
        elif x in phone_num_list:
            new_column_name.append(phone_num_list[0])
        elif x in address_list:
            new_column_name.append(address_list[0])
        elif x in goods_list:
            new_column_name.append(goods_list[0])
        elif x in quantity_list:
            new_column_name.append(quantity_list[0])
        elif x in shipper_list:
            new_column_name.append(shipper_list[0])
        elif x in shipper_num_list:
            new_column_name.append(shipper_num_list[0])
        else:
            new_column_name.append(x)
            unknown_column_name.append(x)
    logging.debug('新列名列表：%s' % new_column_name)
    logging.warning('未识别列名列表：%s' % unknown_column_name)
    df.columns = pd.Index(new_column_name)

    goods = df['商品']
    quantity = df['数量']
    n = len(goods)
    logging.info('订单数量：%i' % n)

    if quantity.apply(lambda x: isinstance(x, int) and (x > 0)).all():
        pass
    else:
        logging.critical('Error: 数量必须为正整数')
        exit()

    gcolumn = []
    for i in range(n):
        glist = str(goods[i]).split('，')
        yun_glist = []
        logging.debug(glist)
        for j in range(len(glist)):
            gitem = glist[j].strip()
            match1 = re.match(r'斯亚旦发皂(.*)块装', gitem)
            match2 = re.match(r'斯亚旦发皂(\d+)箱（每箱20块）', gitem)
            silu_match1 = re.match(r'\(1\)伊花集斯亚旦精油液体手工皂\[300ml]', gitem)
            silu_match2 = re.match(r'\(1\)伊花集斯亚旦精油液体手工皂\[【带1个起泡瓶】300ml]', gitem)
            silu_match3 = re.match(r'\(1\)伊花集斯亚旦精油液体手工皂\[300ml\*(\d+)瓶]', gitem)
            silu_match4 = re.match(r'\(1\)伊花集斯亚旦精油液体手工皂\[【带1个起泡瓶】300ml\*(\d+)瓶]', gitem)
            silu_match5 = re.match(r'\(1\)伊花集斯亚旦精油发皂组合\[基础款100g\+液体皂300ml]', gitem)
            silu_match6 = re.match(r'\(1\)伊花集斯亚旦精油发皂组合\[清爽款100g\+液体皂300ml]', gitem)
            silu_match7 = re.match(r'\(1\)伊花集斯亚旦精油发皂组合\[基础款100g\+液体皂300ml\+起泡瓶1个]', gitem)
            silu_match8 = re.match(r'\(1\)伊花集斯亚旦精油发皂组合\[清爽款100g\+液体皂300ml\+起泡瓶1个]', gitem)
            silu_match9 = re.match(r'\(1\)伊花集斯亚旦植物发皂\[基础款100g]', gitem)
            silu_match10 = re.match(r'\(1\)伊花集斯亚旦植物发皂\[基础款100g\*(\d+)盒]', gitem)
            silu_match11 = re.match(r'\(1\)伊花集斯亚旦植物发皂\[清爽款100g]', gitem)
            silu_match12 = re.match(r'\(1\)伊花集斯亚旦植物发皂\[清爽款100g\*(\d+)盒]', gitem)
            silu_match13 = re.match(r'\(1\)伊花集斯亚旦植物发皂\[基础款100g\+清爽款100g]', gitem)
            silu_match14 = re.match(r'\(1\)伊花集斯亚旦草本植物洗发皂礼盒\[4盒洗发皂]', gitem)
            silu_match15 = re.match(r'\(1\)伊花集斯亚旦植物养发系列礼盒\[发油1盒\+发皂2盒]', gitem)
            silu_match16 = re.match(r'\(1\)伊花集冷制驼乳玫瑰芳华洗脸皂\[80g]', gitem)
            silu_match17 = re.match(r'\(1\)伊花集冷制驼乳玫瑰芳华洗脸皂\[80g\*(\d+)盒]', gitem)
            silu_match18 = re.match(r'\(1\)伊花集奥斯曼植物发油\[10ml]', gitem)
            silu_match19 = re.match(r'\(1\)伊花集奥斯曼植物发油\[10ml\*(\d+)盒]', gitem)
            silu_match20 = re.match(r'\(1\)伊花集奥斯曼植物发油\[【组合装】奥斯曼发油10ml\+基础款发皂100g]', gitem)
            silu_match21 = re.match(r'\(1\)伊花集奥斯曼植物发油\[【组合装】奥斯曼发油10ml\+清爽款发皂100g]', gitem)
            if match1:
                quant = re.sub('两', '二', match1.groups()[0])
                quant = cn2an.cn2an(quant)
                yun_glist.append([golden[0], quant * quantity[i]])
            elif match2:
                quant = int(match2.groups()[0])
                yun_glist.append([golden[0], quant * 20 * quantity[i]])
            elif silu_match1:
                yun_glist.append([golden[8], quantity[i]])
            elif silu_match2:
                yun_glist.append([golden[8], quantity[i]])
                yun_glist.append([golden[7], quantity[i]])
            elif silu_match3:
                quant = int(silu_match3.groups()[0])
                yun_glist.append([golden[8], quant * quantity[i]])
            elif silu_match4:
                quant = int(silu_match4.groups()[0])
                yun_glist.append([golden[8], quant * quantity[i]])
                yun_glist.append([golden[7], quantity[i]])
            elif silu_match5:
                yun_glist.append([golden[0], quantity[i]])
                yun_glist.append([golden[8], quantity[i]])
            elif silu_match6:
                yun_glist.append([golden[1], quantity[i]])
                yun_glist.append([golden[8], quantity[i]])
            elif silu_match7:
                yun_glist.append([golden[0], quantity[i]])
                yun_glist.append([golden[8], quantity[i]])
                yun_glist.append([golden[7], quantity[i]])
            elif silu_match8:
                yun_glist.append([golden[1], quantity[i]])
                yun_glist.append([golden[8], quantity[i]])
                yun_glist.append([golden[7], quantity[i]])
            elif silu_match9:
                yun_glist.append([golden[0], quantity[i]])
            elif silu_match10:
                quant = int(silu_match10.groups()[0])
                yun_glist.append([golden[0], quant * quantity[i]])
            elif silu_match11:
                yun_glist.append([golden[1], quantity[i]])
            elif silu_match12:
                quant = int(silu_match12.groups()[0])
                yun_glist.append([golden[1], quant * quantity[i]])
            elif silu_match13:
                yun_glist.append([golden[0], quantity[i]])
                yun_glist.append([golden[1], quantity[i]])
            elif silu_match14:
                yun_glist.append([golden[0], 4 * quantity[i]])
                yun_glist.append([golden[6], quantity[i]])
                yun_glist.append([golden[10], quantity[i]])
            elif silu_match15:
                yun_glist.append([golden[2], quantity[i]])
                yun_glist.append([golden[0], 2 * quantity[i]])
                yun_glist.append([golden[5], quantity[i]])
                yun_glist.append([golden[10], quantity[i]])
            elif silu_match16:
                yun_glist.append([golden[4], quantity[i]])
            elif silu_match17:
                quant = int(silu_match17.groups()[0])
                yun_glist.append([golden[4], quant * quantity[i]])
            elif silu_match18:
                yun_glist.append([golden[2], quantity[i]])
            elif silu_match19:
                quant = int(silu_match19.groups()[0])
                yun_glist.append([golden[2], quant * quantity[i]])
            elif silu_match20:
                yun_glist.append([golden[2], quantity[i]])
                yun_glist.append([golden[0], quantity[i]])
            elif silu_match21:
                yun_glist.append([golden[2], quantity[i]])
                yun_glist.append([golden[1], quantity[i]])
            else:
                logging.error('未识别的商品：%s' % glist[j])
                exit()
        logging.debug(yun_glist)
        gcell = ''
        for g in yun_glist:
            if g[0] == golden[0] or g[0] == golden[1] or g[0] == golden[8]:
                yun_glist.append([golden[14], 1])
                break
        for j in range(len(yun_glist)):
            if j == 0:
                gcell = '%s*%s' % (yun_glist[j][0], yun_glist[j][1])
            else:
                gcell += '，%s*%s' % (yun_glist[j][0], yun_glist[j][1])
        gcolumn.append(gcell)
    logging.debug(gcolumn)

    df['商品'] = gcolumn
    df['数量'] = 1

    # 获取文件名（包含扩展名）
    file_name = os.path.basename(excel_path)

    # 获取文件名（不包含扩展名）
    (file_name_without_ext, file_ext) = os.path.splitext(file_name)
    output_dir = 'output'

    # 检查目录是否存在
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    output_file_path = '%s/%s-云仓%s' % (output_dir, file_name_without_ext, file_ext)
    with pd.ExcelWriter(output_file_path) as writer:
        df.to_excel(writer, engine='openpyxl', index=False)
        # 获取工作表对象
        worksheet = writer.sheets['Sheet1']

        # 创建字体样式
        header_font = Font(name='黑体', size=14)
        body_font = Font(name='宋体', size=12)

        # 创建对齐方式
        center_alignment = Alignment(horizontal='center', vertical='center')

        # 创建填充样式
        header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

        # 设置表头样式
        for cell in worksheet[1]:
            cell.font = header_font
            cell.alignment = center_alignment
            cell.fill = header_fill

        # 设置数据行样式
        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                cell.font = body_font
                cell.alignment = center_alignment
        # 遍历所有列
        for col_idx, column in enumerate(df):
            # 获取列的最大长度
            column_length = max(df[column].astype(str).map(len).max(), len(column))
            # 将列索引转换为字母
            col_letter = get_column_letter(col_idx + 1)
            # 设置列宽
            worksheet.column_dimensions[col_letter].width = column_length + 8
    logging.info('%s reformed successfully.%s' % (file_name_without_ext, os.linesep))


if __name__ == '__main__':
    ordertrans('input/丝路228.xlsx')
